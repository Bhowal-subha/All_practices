import openpyxl

get_book = openpyxl.load_workbook(r"C:\Users\subhowal\Downloads\example.xlsx")
# This is for particular one sheet which is active
# bk = get_book.active

# by sheet name we can choose
sheet_name = "Sheet2"
bk = get_book[sheet_name]

# particular row and cloumn value we can select
a = bk.cell(row=2, column=3)
print(a.value)

# To write data in excel
# bk.cell(row=7,column=2).value="q"
# get_book.save(r"C:\Users\subhowal\Downloads\example.xlsx")
# print(bk.cell(row=7,column=2).value)

#write from a dic
# dic={'cell2': 't', 'cell3': 'u', 'cell4': 'v'}
# for k,v in dic.items():
#     for i in range(1,bk.max_row+1):
#         if bk.cell(row=i,column=1)=="tC3":
#             for j in range(2,bk.max_column+1):
#                 bk.cell(row=i,column=j).value=dic[v]
#
#
# # Get max no. of row and column
# print(bk.max_row)
# print(bk.max_column)
#
# # Extract all value
# # for i in range(1,bk.max_row+1):
# #     for j in range(1,bk.max_column+1):
# #         print(bk.cell(row=i,column=j).value)
#
# Extract for particular section
dict = {}
for i in range(1, bk.max_row + 1):
    if bk.cell(row=i, column=1).value == "tC2":
        for j in range(2, bk.max_column + 1):
            dict[bk.cell(row=1, column=j).value] = bk.cell(row=i, column=j).value
print([dict])

# Store in a Dict
# dic = {}
# for i in range(1, bk.max_row + 1):
#     if i == 3:
#         for j in range(2, bk.max_column + 1):
#             dic[bk.cell(row=1, column=j).value] = bk.cell(row=i, column=j).value
# print([dic])

# Copy data from one sheet to another
s_sheet = "Sheet1"
sk = get_book[s_sheet]
d_sheet = "Sheet3"
dk = get_book[d_sheet]

for i in sk.iter_rows():
    for j in i:
        dk[j.coordinate].value = j.value
get_book.save(r"C:\Users\subhowal\Downloads\example.xlsx")
